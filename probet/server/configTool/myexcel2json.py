#coding:utf-8
import os;
import sys;
import time;
from datetime import datetime
import warnings;
import openpyxl;
import openpyxl.cell;
import cProfile;
import colorama
from colorama import Fore, Back, Style

import re;

def LogInfo(strText):
	sys.stderr.write(Fore.GREEN + Style.BRIGHT + strText + Fore.RESET + Style.RESET_ALL);
def LogError(strText):
	sys.stderr.write(Fore.RED + Style.BRIGHT + strText + Fore.RESET + Style.RESET_ALL);

# 是否是float, 简单实现
def StringIsFloat(strText):
	iPeriodCount = 0;
	if(strText is None or strText == ''): return False;
	for index in range(0, len(strText)):
		i = strText[index];
		if(index == 0 and i == '-'): continue;
		if(i.isdigit() == True): continue;

		if(i != '.'): return False;

		iPeriodCount = iPeriodCount + 1;
		if(iPeriodCount > 1.1): return False;

	return True;

# lua 转义
def StringLuaEscape(strText):
	strEscape = strText;

	strEscape = strEscape.replace("\\", r"\\");

	strEscape = strEscape.replace("\a", r"\a");
	strEscape = strEscape.replace("\b", r"\b");
	strEscape = strEscape.replace("\f", r"\f");
	strEscape = strEscape.replace("\n", r"\n");
	strEscape = strEscape.replace("\r", r"\r");
	strEscape = strEscape.replace("\t", r"\t");
	strEscape = strEscape.replace("\v", r"\v");
	strEscape = strEscape.replace('"',  r'\"');
	strEscape = strEscape.replace("'",  r"\'");
	return strEscape;

# json 转义 http://json.org/
def StringJsonEscape(strText):
	strEscape = strText;

	strEscape = strEscape.replace("\\", r"\\");

	strEscape = strEscape.replace('"',  r'\"');
    #strEscape = strEscape.replace('/',  r'\/');
	strEscape = strEscape.replace("\b", r"\b");
	strEscape = strEscape.replace("\f", r"\f");
	strEscape = strEscape.replace("\n", r"\n");
	strEscape = strEscape.replace("\r", r"\r");
	strEscape = strEscape.replace("\t", r"\t");

	return strEscape;

# 导出 lua 格式
def DumpSheetToLua(strSrcExcel, oExcelSheet, iRealRowCount, iColCount, oListHeaderName, oListHeaderType, oAllRowList):

	strExcelBaseName = os.path.basename(strSrcExcel)
	strExcelBaseName = os.path.splitext(strExcelBaseName)[0];

	strRet = "\r\n\r\n" + "local t = {}; \r\n";
	strRet = strRet + "\r\n" + 't["name"] = "{0}.{1}";'.format(strExcelBaseName, oExcelSheet.title);
	strRet = strRet + "\r\n" + 't["rownum"] = {0};'.format(iRealRowCount);
	strRet = strRet + "\r\n" + 't["colnum"] = {0};'.format( len(oListHeaderName) );

	strKeyListLine = 't["keylist"] = { ';
	strKeyIndexLine = 't["keyindex"] = { ';

	# key 名和索引
	for i in range(0, len(oListHeaderName)):
		strKeyName = oListHeaderName[i];
		strKeyListLine = strKeyListLine + '"{0}", '.format(strKeyName);
		strKeyIndexLine = strKeyIndexLine + '["{0}"] = {1}, '.format(strKeyName, i + 1);

	strRet = strRet + "\r\n" + strKeyListLine + "};";
	strRet = strRet + "\r\n" + strKeyIndexLine + "};";

	# 所有行
	strAllKeyListText = 't["k"] = { ';
	strAllRowListText = 't["r"] = { ';
	for i in range(0, len(oAllRowList)):
		oThisRow = oAllRowList[i];
		strThisKeyLine = '"{0}", '.format(oThisRow[0]);
		strThisRowLine = ' ["{0}"] = {{ '.format(oThisRow[0]);
		for j in range(0, len(oThisRow)):
			strThisType = oListHeaderType[j];
			strEncompass = '"';
			#if(strThisType == "number"): strEncompass = '';
			oThisValue = oThisRow[j];
			strLuaEscaped = oThisValue;
			if(StringIsFloat(oThisValue)):
				strEncompass = '';
			else:
				strLuaEscaped = StringLuaEscape(oThisValue);
			strThisRowLine = strThisRowLine + '{0}{1}{2}, '.format(strEncompass, strLuaEscaped, strEncompass);
		strThisRowLine = strThisRowLine + '};';

		strAllKeyListText = strAllKeyListText + strThisKeyLine;
		strAllRowListText = strAllRowListText + "\r\n" + strThisRowLine;

	strAllKeyListText = strAllKeyListText + "}; \r\n";
	strAllRowListText = strAllRowListText + "\r\n}; \r\n";

	strRet = strRet + "\r\n\r\n" + strAllKeyListText;
	strRet = strRet + "\r\n\r\n" + strAllRowListText;

	strRet = strRet + "\r\n\r\n" + "return t; \r\n\r\n";

	return strRet;


# 导出 json 格式
def DumpSheetToJson(strSrcExcel, oExcelSheet, iRealRowCount, iColCount, oListHeaderName, oListHeaderType, oAllRowList,oListHeaderCS,strJsonForWho='cs'):
	strRet = "{ \r\n";
	iRowLen = len(oAllRowList);
	for i in range(0, iRowLen):
		oThisRow = oAllRowList[i];
		strThisRowLine = '    "{0}": {{\r\n'.format(oThisRow[0]);
		iColLen = len(oThisRow);
		for j in range(0, iColLen):
			strThisType = oListHeaderType[j];
			cSHeader = oListHeaderCS[j];
			oThisValue = oThisRow[j];

			if cSHeader.find(strJsonForWho) < 0:
				continue;

			#默认是字符串处理
			strBeginEncompass = '"';
			strEndEncompass = '"';
			if(strThisType == "number"):
				strBeginEncompass = '';
				strEndEncompass = '';
				strJsonEscaped = oThisValue;
			elif(strThisType == "timestamp"):
				strBeginEncompass = '';
				strEndEncompass = '';
				strJsonEscaped = int(datetime.strptime(oThisValue, "%Y-%m-%d %H:%M:%S").timestamp());
			elif(strThisType == "string"):
				#使用默认的

				strJsonEscaped = StringJsonEscape(oThisValue);
				pass
			elif(strThisType == "rn_require"):
				strBeginEncompass = "require('";
				strEndEncompass = "')";

				strJsonEscaped = StringJsonEscape(oThisValue);


			if(j == iColLen - 1):
				strThisRowLine = strThisRowLine + '        "{0}": {1}{2}{3}\r\n'.format(oListHeaderName[j], strBeginEncompass, strJsonEscaped, strEndEncompass);
			else:
				strThisRowLine = strThisRowLine + '        "{0}": {1}{2}{3},\r\n'.format(oListHeaderName[j], strBeginEncompass, strJsonEscaped, strEndEncompass);
		# 行尾了
		if(i == iRowLen - 1):
			strThisRowLine = strThisRowLine + '    }\r\n';
		else:
			strThisRowLine = strThisRowLine + '    },\r\n';
		strRet = strRet + strThisRowLine;

	strRet = strRet + "}";
	return strRet;


# 单元合法性检测
def CheckColumnValid(iColIndex, strColText, oListHeaderName, oListHeaderType, oListHeaderPattern, oListHeaderRegex):
	#if( oListHeaderName[iColIndex] == "huixin_rate"):
	#	pass;
	if(strColText is None): return u"非字符串类型不能为空";

	strColText = str(strColText);
	strType = oListHeaderType[iColIndex];
	if(strType == "number"):
		if(StringIsFloat(strColText) == False):
			return u"非法的数字类型";
		else:
			return "";
		pass;
	elif(strType == "string"):
		return "";
	elif(strType == "numberlist"):
		strSep = oListHeaderPattern[iColIndex];
		arrList = strColText.split(strSep);
		iLen = len(arrList);
		if(iLen <= 0): return "非法的数字列表 sep[{0}]".format(strSep);
		for x in arrList:
			if(StringIsFloat(x) == False):
				return "非法的数字列表 sep[{0}]".format(strSep);
		return "";
	elif(strType == "stringlist"):
		return "";
	elif(strType == "regex"):
		strPattern = oListHeaderPattern[iColIndex];
		if( strPattern == ""): return "";
		oRegex = oListHeaderRegex[iColIndex];
		oRegReturn = oRegex.search(strColText);
		if(oRegReturn is None):
			return "非法的正则内容 regex[{0}]".format(strPattern);

	elif(strType == "combo"):
		# + 号分隔的字段，第一个是大分隔符，第二个是大分隔数量, 第三个是小分隔符，第4个是小分隔数量
		strPattern = oListHeaderPattern[iColIndex];
		arrPatternArray = strPattern.split("!");
		if(arrPatternArray is None or len(arrPatternArray) < 4):
			return "非法的combo模式头 [{0}]".format(strPattern);

		strBigSep = arrPatternArray[0];
		iBigCount = int(arrPatternArray[1]);
		strSmallSep = arrPatternArray[2];
		iSmallCount = int(arrPatternArray[3]);

		arrBigArrayList = strColText.split(strBigSep);
		if(arrBigArrayList is None or len(arrBigArrayList) != iBigCount):
			return "非法的combo内容, 1维数量不对, combo[{0}]".format(strPattern);

		for i in range(0, len(arrBigArrayList)):
			strThisItem = arrBigArrayList[i];
			arrThisArray = strThisItem.split(strSmallSep);
			if(arrThisArray is None or len(arrThisArray) != iSmallCount):
				return "非法的combo内容, 2维数量不对, 子数组[{0}]".format(strThisItem);
	elif(strType == 'rn_require'):
		return "";
	elif (strType == 'timestamp'):
		return "";
	else:
		return "非法的表格单元类型: [{0}]".format(strType);
	return "";


# 导出一个 excel 表单
def MyExportSheetToDir(strSrcExcel:str, strDstDirJson:str, oExcelSheet):
	# 检查最小行数，列数
	iRowCount = oExcelSheet.max_row;
	iColCount = oExcelSheet.max_column;
	if(iRowCount < 4): # 最少要3行, 标题，名字，cs:valildator
		LogError("\r\nERROR: 表行数不够:{0}".format(iRowCount));
		return -4;
	if iColCount < 2: # 最少要2列, ID, xxx
		LogError("\r\nERROR: 表列数不够:{0}".format(iColCount));
		return -4;

	oListHeaderTitle = []; # 标题
	oListHeaderName = []; # 列名
	oListHeaderCS = []; # 是否是 cs
	oListHeaderType = []; # 类型
	oListHeaderPattern = []; # 数组分隔号或模式
	oListHeaderRegex = []; # 正则模式, 编译后的

	iMaxColCount = iColCount;
	oExcelRowList = [];

	for var in oExcelSheet.rows:
		oExcelRowList.append(var)

	oExcelRow_0 = oExcelRowList[0];
	oExcelRow_1 = oExcelRowList[1];
	oExcelRow_2 = oExcelRowList[2];
	oExcelRow_3 = oExcelRowList[3];


	# 处理头
	for i in range(0, iMaxColCount):
		if(oExcelRow_1[i].value is None):
			iColCount = i;
			break;

		strThisTitle = str(oExcelRow_0[i].value);
		strThisName = str(oExcelRow_1[i].value);
		strThisCSValidator = str(oExcelRow_2[i].value);

		if(strThisName == ""):
			iColCount = i;
			break;

		if( (oExcelRow_0[i].value is None) or (strThisTitle == "") ):
			LogError("\r\nERROR: 标题不能为空: 表单[{0}], 第[{1}]列 \r\n".format(oExcelSheet.title, i));
			return -5;

		if( (oExcelRow_1[i].value is None) or (strThisName == "") ):
			LogError("\r\nERROR: 名称不能为空: 表单[{0}], 第[{1}]列 \r\n".format(oExcelSheet.title, i));
			return -5;

		if( (oExcelRow_2[i].value is None) or (strThisCSValidator == "") ):
			LogError("\r\nERROR: 导出选项不能为空: 表单[{0}], 第[{1}]列 \r\n".format(oExcelSheet.title, i));
			return -5;

		strThisNameLower = strThisName.lower();
		if(strThisNameLower == "item" or strThisNameLower == "count" or strThisNameLower == "containskey"):
			LogError("\r\nERROR: 列名不能是 item/count/containskey: 表单[{0}], 第[{1}]列 \r\n".format(oExcelSheet.title, i));
			return -5;

		# separator by : CS:number/string/numberlist/stringlist/regex:sep/pattern/rn_require/timestamp
		arrCSValidator = strThisCSValidator.split(':');

		# ~ 开头的CS列是注释
		if( (arrCSValidator is not None) and (len(arrCSValidator) >= 1) and (arrCSValidator[0] == "~") ):
			if(strThisName != ""):
				LogError("\r\nERROR: 注释列必须是最后一列，而且名字要留空: 表单[{0}], 第[{1}]列 \r\n".format(oExcelSheet.title, i));
				return -5;
			iColCount = iColCount - 1;
			continue;

		if(len(arrCSValidator) > 3):
			strTemp = arrCSValidator[2];
			for i in range(3, len(arrCSValidator)):
				strTemp = strTemp + ':' + arrCSValidator[i];
			arrCSValidator[2] = strTemp;

		strCSField = arrCSValidator[0].lower();
		if( not (strCSField == "cs" or strCSField == "c" or strCSField == "s") ):
			LogError("\r\nERROR: 导出选项要以 cs/c/s开头: 第[{0}]列 \r\n".format(i));
			return -5;
		oListHeaderCS.append(strCSField);

		oListHeaderTitle.append(strThisTitle);
		oListHeaderName.append(strThisName);

		strType = "string";

		if(len(arrCSValidator) >= 2):
			strThisType = arrCSValidator[1].lower();
			if(strThisType == "number" or strThisType == "string" or strThisType == "numberlist" or strThisType == "stringlist" or strThisType == "regex" or strThisType == "combo" or strThisType == "timestamp"):
				strType = strThisType;
			elif strThisType == "rn_require":
				strType = "rn_require";
			else:
				LogError("\r\nERROR: 非法的表格单元类型: 第[{0}]列, 类型[{1}]\r\n".format(i, strThisType));
				return -5;

		strPattern = "";
		if(len(arrCSValidator) >= 3):
			strPattern = arrCSValidator[2];
		# 正则模式
		oRegex = strPattern;
		if(strPattern != ""):
			oRegex = re.compile(strPattern);
		oListHeaderType.append(strType);

		oListHeaderPattern.append(strPattern);
		oListHeaderRegex.append(oRegex);
	#print('excel 行数[{0}], 列数[{1}]'.format(iRowCount, iColCount));

	# 首列必须是id
	if(oListHeaderName[0] != "id"):
		LogError("\r\nERROR: 第一列必须是ID: 表单[{0}], [{1}] \r\n".format(oExcelSheet.title, oListHeaderName[0]));
		return -5;

	# 处理所有行
	oAllRowList = [];
	for i in range(3, iRowCount):
		oRow = oExcelRowList[i];
		oCell0 = oRow[0]
		strID = oCell0.value;
		strIDType = oCell0.data_type;

		if(strID is None): break;
		strID = str(strID);
		#if(strID.isdigit() == False):
		#	sys.stderr.write("ERROR: 第一列ID必须是整数: row[{0}]=[{1}] \r\n".format(i, strID));
		#	return -5;
		oTempRow = [];
		# ID
		oTempRow.append( strID );

		strIDLower = strID.lower();
		if(strIDLower == "item" or strIDLower == "count" or strIDLower == "containskey"):
			LogError("\r\nERROR: ID 不能是 item/count/containskey: 表单[{0}], 第[{1}]行 \r\n".format(oExcelSheet.title, i));
			return -5;
			


		# 其它
		for j in range(1, iColCount):
			oThisCell = oRow[j];
			strColText = oThisCell.value;

			if(strColText is None and oListHeaderType[j] == "string"):
				# 默认值
				strThisColValue = "";
			else:
				# 合法性检测
				strRet = CheckColumnValid(j, strColText, oListHeaderName, oListHeaderType, oListHeaderPattern, oListHeaderRegex);
				if(strRet != ""):
					LogError("\r\nERROR: 表格单元不合法 row[{0}] col[{1}] cell[{2}] ex[{3}]\r\n".format(i, j, strColText, strRet));
					return -5;
				strThisColValue = strColText;


			oTempRow.append( str(strThisColValue) );
		# 加入行
		oAllRowList.append(oTempRow);



	strExcelBaseName = os.path.basename(strSrcExcel)
	strExcelBaseName = os.path.splitext(strExcelBaseName)[0];

	# 导成 json 格式字符串
	strJsonDump = DumpSheetToJson(strSrcExcel, oExcelSheet, iRowCount - 3, iColCount, oListHeaderName, oListHeaderType, oAllRowList,oListHeaderCS,'s');

	# 调试打印出来看看
	#print("jsonfile[{0}]".format(strJsonDump));
	#strLuaFileName = '{0}.{1}.txt'.format(strExcelBaseName, oExcelSheet.title);
	strLuaFileName = '{0}.py'.format(strExcelBaseName);
	oFileJson = open(strDstDirJson + strLuaFileName, "wb");
	oFileJson.write("{0}_cfg=".format(strExcelBaseName,oExcelSheet.title).encode('utf-8'));
	oFileJson.write(strJsonDump.encode('utf-8'));
	oFileJson.close();

	'''
	strJsonDump = DumpSheetToJson(strSrcExcel, oExcelSheet, iRowCount - 3, iColCount, oListHeaderName, oListHeaderType, oAllRowList,oListHeaderCS,'c');
	strLuaFileName = '{0}.js'.format(strExcelBaseName);
	oFileJs = open(strDstDirJson  + strLuaFileName, "wb");
	oFileJs.write("global.{0}_cfg=\n".format(strExcelBaseName,oExcelSheet.title).encode('utf-8'));
	oFileJs.write(strJsonDump.encode('utf-8'));
	oFileJs.close();
	'''
	
	return 0;

# 导出 excel 到指定目录
def MyExportExcelToDir(strSrcExcel:str, strDstDirJson:str):
	if(os.path.isfile(strSrcExcel) == False):
		LogError("\r\nERROR: 找不到excel文件: [{0}] \r\n".format(strSrcExcel));
		return -2;
	if(os.path.isdir(strDstDirJson) == False):
		LogError("\r\nERROR: 目标目录不存在: [{0}] \r\n".format(strDstDirJson));
		return -3;

	warnings.filterwarnings("ignore");

	strBaseName = os.path.basename(strSrcExcel);

	tStartTime = time.time();
	LogInfo("导出配表中 [{0}] ... ".format(strBaseName));

	# 打开 excel 文件
	oWorkbook = openpyxl.load_workbook(strSrcExcel);

	for oExcelSheet in oWorkbook.worksheets:
		if(oExcelSheet.title[0] == '~' or oExcelSheet.title.lower() == 'readme'):
			continue;
		# 导出这个 sheet
		iRet = MyExportSheetToDir(strSrcExcel, strDstDirJson, oExcelSheet);
		if(iRet != 0): return iRet;

	tEndTime = time.time();
	LogInfo(u"完成({0:.2f}秒)\r\n".format( float((tEndTime - tStartTime)), strBaseName));

	return 0;

#
def main(oListArgv):
	colorama.init();
	strSourceExcel = oListArgv[1];
	strDstDirJson = oListArgv[2];
	iRet = MyExportExcelToDir(strSourceExcel, strDstDirJson);
	return iRet;

# for module
if __name__ == "__main__":

	oListArg = sys.argv;
	strCurrentPath = os.path.dirname(os.path.realpath(__file__));
	print(strCurrentPath)
	if(False):
		#strRootDir = r'D:\ReactNative\trunk\tool';
		strRootDir = strCurrentPath
		oListArg = [ sys.argv[0], strRootDir + r'/StoreConfig_all.xlsx', strRootDir];

	if(len(oListArg) < 3):
		LogError("Usage: myexcel2json.py src.xlsx diroutjson");
		sys.exit(-2);

	if(False):
		cProfile.run("iRet = main(oListArg);sys.exit(iRet);", sort = 2);
	else:
		iRet = main(oListArg);
		sys.exit(iRet);


